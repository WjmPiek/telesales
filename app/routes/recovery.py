from datetime import date, timedelta, datetime
import secrets
import json
import os
import re
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file, abort, jsonify
from flask_login import login_required, current_user
from openpyxl import load_workbook
from app import db
from app.models import LapsedPolicy, RecoveryCallLog, ClientApplication, PolicyProduct, TelesalesScriptSession, ApplicationSignature, ClientFicaDocument, AuditLog
from app.security import permission_required
from app.services.pdf_service import generate_telesales_script_pdf, generate_application_pdf, generate_popia_pdf, generate_disclosure_pdf, generate_fica_pdf
from app.services.email_service import send_email
from app.services.whatsapp_service import send_whatsapp_message
from app.services.compliance_service import dob_from_sa_id, age_from_dob, classify_product_template, assert_application_rules
from app.services.branch_access import scope_by_branch, ensure_branch_access, can_view_all_branches, is_branch_manager, user_branch

recovery_bp = Blueprint("recovery", __name__, url_prefix="/recovery")

LEAD_OPEN_STATUSES = [
    "New",
    "Imported",
    "Called",
    "No Answer",
    "Callback",
    "Interested",
    "Application Started",
    "Signature Sent",
    "FICA Outstanding",
    "QA Review",
]
LEAD_CLOSED_STATUSES = ["Approved", "Rejected", "Closed", "Reinstated", "Suspense"]
SUSPENSE_STATUS = "Suspense"

ID_HEADERS = ["ID_Number", "ID Number", "IDNumber", "ID No", "ID_No", "ID", "SA ID", "SA_ID", "Identity Number", "Client ID Number", "IdentityNumber", "IdNumber", "RSA ID"]
CONTACT_HEADERS = ["Cell_Number", "Cell Number", "Cell", "Mobile", "Mobile Number", "Phone", "Contact Number", "Contact_Number", "home_tel", "Home Tel", "Telephone", "Tel", "Client Cell", "Client Mobile"]
EMAIL_HEADERS = ["Email", "Email Address", "Email_Address", "E-mail", "E Mail", "Client Email", "EmailAddress"]
COMPANY_HEADERS = ["Company", "Company Name", "Company_Name", "Franchise", "Business Client", "Business Client Name", "Client Company", "Employer", "CollectedatBranch", "Collected at Branch", "Branch"]


def _clean_import_value(value):
    if value is None:
        return ""
    txt = str(value).strip()
    if txt.lower() in {"none", "nan", "null", "n/a", "na", "-"}:
        return ""
    return txt


def _norm_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _row_value(data, names):
    # Handles exact headers, spacing/case changes, punctuation changes and common import variants.
    normalised = {_norm_header(k): v for k, v in data.items()}
    for name in names:
        if name in data and _clean_import_value(data.get(name)):
            return _clean_import_value(data.get(name))
        key = _norm_header(name)
        if key in normalised and _clean_import_value(normalised.get(key)):
            return _clean_import_value(normalised.get(key))
    return ""


def _is_missing(value):
    txt = _clean_import_value(value)
    if not txt:
        return True
    return txt.lower() in {"0", "000", "0000000000000", "unknown", "not available", "missing", "no email", "no phone", "no cell"}


def _missing_contact_fields(data):
    missing = []
    id_number = _row_value(data, ID_HEADERS)
    contact_number = _row_value(data, CONTACT_HEADERS)
    email_address = _row_value(data, EMAIL_HEADERS)
    id_digits = re.sub(r"\D", "", id_number or "")
    contact_digits = re.sub(r"\D", "", contact_number or "")
    if _is_missing(id_number) or not id_digits:
        missing.append("ID number")
    if _is_missing(contact_number) or not contact_digits:
        missing.append("contact number")
    if _is_missing(email_address) or "@" not in str(email_address):
        missing.append("email address")
    return missing, id_number, contact_number, email_address


def _append_comment(existing, note):
    existing = (existing or "").strip()
    return (existing + "\n" + note).strip() if existing else note



def _policy_missing_fields(policy):
    missing = []
    if _is_missing(getattr(policy, "id_number", None)):
        missing.append("ID number")
    if _is_missing(getattr(policy, "cell_number", None)) and _is_missing(getattr(policy, "home_tel", None)):
        missing.append("contact number")
    if _is_missing(getattr(policy, "email_address", None)) or "@" not in str(getattr(policy, "email_address", "") or ""):
        missing.append("email address")
    return missing


def _move_missing_policies_to_suspense(query=None):
    query = query or LapsedPolicy.query
    moved = 0
    for p in query.filter(LapsedPolicy.recovery_status != SUSPENSE_STATUS).all():
        missing = _policy_missing_fields(p)
        if missing:
            p.recovery_status = SUSPENSE_STATUS
            p.assigned_agent_id = None
            p.next_action_date = None
            p.suspense_reason = ", ".join(missing)
            if not getattr(p, "company_name", None):
                p.company_name = p.franchise or p.branch
            p.comments = _append_comment(p.comments, "SUSPENSE: Missing " + ", ".join(missing) + ". Client cannot be contacted until business client/branch fixes these policy details.")
            moved += 1
    if moved:
        db.session.commit()
    return moved

OUTCOME_TO_STATUS = {
    "No Answer": "No Answer",
    "Voicemail": "No Answer",
    "Wrong Number": "Closed",
    "Number Invalid": "Closed",
    "Not Interested": "Closed",
    "Callback Requested": "Callback",
    "Interested": "Interested",
    "Application Started": "Application Started",
    "Signature Sent": "Signature Sent",
    "FICA Outstanding": "FICA Outstanding",
    "QA Review": "QA Review",
    "Approved": "Approved",
    "Rejected": "Rejected",
    "Dispute / Complaint": "QA Review",
    "Reinstated": "Reinstated",
    "Wants Reinstatement": "Application Started",
    "Wants New Policy": "Application Started",
}

CALL_OUTCOMES = list(OUTCOME_TO_STATUS.keys())
CALLBACK_OUTCOMES = {"No Answer", "Voicemail", "Callback Requested"}


def open_recovery_query():
    return scope_by_branch(
        LapsedPolicy.query.filter(LapsedPolicy.recovery_status.notin_(LEAD_CLOSED_STATUSES)),
        LapsedPolicy,
        agent_col=LapsedPolicy.assigned_agent_id,
    )


@recovery_bp.route("/")
@login_required
@permission_required("recovery.view")
def queue():
    status = request.args.get("status")
    query = open_recovery_query()
    if status:
        query = query.filter(LapsedPolicy.recovery_status == status)
    policies = query.order_by(
        LapsedPolicy.next_action_date.asc().nullslast(),
        LapsedPolicy.imported_at.asc()
    ).limit(200).all()

    status_counts = {
        row[0] or "New": row[1]
        for row in open_recovery_query().with_entities(LapsedPolicy.recovery_status, db.func.count(LapsedPolicy.id))
        .group_by(LapsedPolicy.recovery_status).all()
    }
    return render_template(
        "recovery/queue.html",
        policies=policies,
        status_counts=status_counts,
        open_statuses=LEAD_OPEN_STATUSES,
        active_status=status,
        today=date.today(),
    )




@recovery_bp.route("/pipeline")
@login_required
@permission_required("recovery.view")
def pipeline():
    """Kanban-style lead pipeline for daily telesales follow-up."""
    columns = [
        "Imported", "No Answer", "Callback", "Interested",
        "Application Started", "Signature Sent", "FICA Outstanding",
        "QA Review", "Approved", "Rejected"
    ]
    cards = {}
    base = scope_by_branch(LapsedPolicy.query, LapsedPolicy, agent_col=LapsedPolicy.assigned_agent_id)
    for status in columns:
        cards[status] = (
            base.filter(LapsedPolicy.recovery_status == status)
            .order_by(LapsedPolicy.next_action_date.asc().nullslast(), LapsedPolicy.imported_at.asc())
            .limit(40)
            .all()
        )
    return render_template("recovery/pipeline.html", columns=columns, cards=cards, today=date.today())


@recovery_bp.route("/<int:policy_id>/status", methods=["POST"])
@login_required
@permission_required("recovery.call")
def update_status(policy_id):
    """Update a lead status from the Kanban board without opening the full call screen."""
    p = LapsedPolicy.query.get_or_404(policy_id)
    ensure_branch_access(p, agent_attr="assigned_agent_id")
    new_status = (request.form.get("status") or "").strip()
    allowed = set(LEAD_OPEN_STATUSES + LEAD_CLOSED_STATUSES + [SUSPENSE_STATUS])
    if new_status not in allowed:
        return jsonify({"ok": False, "error": "Invalid status"}), 400
    old_status = p.recovery_status
    p.recovery_status = new_status
    if new_status == "Callback" and not p.next_action_date:
        p.next_action_date = date.today()
    db.session.add(AuditLog(user_id=current_user.id, action="LEAD_STATUS_CHANGED", entity_type="LapsedPolicy", entity_id=str(p.id), details=f"{old_status} -> {new_status}"))
    db.session.commit()
    return jsonify({"ok": True, "status": new_status})


@recovery_bp.route("/next")
@login_required
@permission_required("recovery.call")
def next_client():
    policy = open_recovery_query().filter(
        (LapsedPolicy.assigned_agent_id == current_user.id) | (LapsedPolicy.assigned_agent_id.is_(None)),
        (LapsedPolicy.next_action_date.is_(None)) | (LapsedPolicy.next_action_date <= date.today())
    ).order_by(
        LapsedPolicy.next_action_date.asc().nullslast(),
        LapsedPolicy.imported_at.asc()
    ).first()
    if not policy:
        flash("No clients are due for calling right now.", "info")
        return redirect(url_for("recovery.queue"))
    if not policy.assigned_agent_id:
        policy.assigned_agent_id = current_user.id
        db.session.commit()
    return redirect(url_for("recovery.log_call", policy_id=policy.id))




@recovery_bp.route("/callbacks")
@login_required
@permission_required("recovery.view")
def callbacks():
    """Phase 2 callback worklist: overdue, today, upcoming and unscheduled follow-ups."""
    today = date.today()
    base = open_recovery_query().filter(LapsedPolicy.recovery_status == "Callback")
    overdue = base.filter(LapsedPolicy.next_action_date < today).order_by(LapsedPolicy.next_action_date.asc()).all()
    due_today = base.filter(LapsedPolicy.next_action_date == today).order_by(LapsedPolicy.imported_at.asc()).all()
    upcoming = base.filter(LapsedPolicy.next_action_date > today).order_by(LapsedPolicy.next_action_date.asc()).limit(200).all()
    unscheduled = base.filter(LapsedPolicy.next_action_date.is_(None)).order_by(LapsedPolicy.imported_at.asc()).limit(200).all()
    return render_template(
        "recovery/callbacks.html",
        overdue=overdue,
        due_today=due_today,
        upcoming=upcoming,
        unscheduled=unscheduled,
        today=today,
    )


@recovery_bp.route("/<int:policy_id>/timeline")
@login_required
@permission_required("recovery.view")
def client_timeline(policy_id):
    """Phase 2 single client timeline built from existing call, script, application, signature and FICA records."""
    p = LapsedPolicy.query.get_or_404(policy_id)
    ensure_branch_access(p, agent_attr="assigned_agent_id")
    calls = RecoveryCallLog.query.filter_by(lapsed_policy_id=p.id).order_by(RecoveryCallLog.created_at.desc()).all()
    applications = ClientApplication.query.filter_by(lapsed_policy_id=p.id).order_by(ClientApplication.created_at.desc()).all()
    sessions = TelesalesScriptSession.query.filter_by(lapsed_policy_id=p.id).order_by(TelesalesScriptSession.created_at.desc()).all()
    app_ids = [a.id for a in applications]
    signatures = ApplicationSignature.query.filter(ApplicationSignature.application_id.in_(app_ids)).order_by(ApplicationSignature.signed_at.desc()).all() if app_ids else []
    fica_docs = ClientFicaDocument.query.filter(ClientFicaDocument.application_id.in_(app_ids)).order_by(ClientFicaDocument.uploaded_at.desc()).all() if app_ids else []

    events = []
    if p.imported_at:
        events.append({"when": p.imported_at, "type": "Lead Imported", "title": f"Policy {p.policy_number} imported", "details": p.comments or ""})
    if p.next_action_date:
        events.append({"when": datetime.combine(p.next_action_date, datetime.min.time()), "type": "Next Action", "title": f"Next action due: {p.next_action_date}", "details": f"Current status: {p.recovery_status}"})

    for call in calls:
        details = call.notes or ""
        if call.next_action_date:
            details = (details + "\n" if details else "") + f"Next action: {call.next_action_date}"
        events.append({"when": call.created_at, "type": "Call", "title": call.outcome, "details": details, "agent": call.agent.name if call.agent else ""})

    for session in sessions:
        title = f"Script {session.status}"
        if session.qa_score is not None:
            title += f" - QA {session.qa_score}%"
        events.append({"when": session.completed_at or session.created_at, "type": "Script", "title": title, "details": session.blocked_reason or session.qa_result or "", "agent": session.agent.name if session.agent else ""})

    for app in applications:
        events.append({"when": app.created_at, "type": "Application", "title": f"Application {app.application_ref} created", "details": f"Status: {app.status}; Type: {app.application_type}; Premium: {app.monthly_premium}"})
        if app.updated_at and app.updated_at != app.created_at:
            events.append({"when": app.updated_at, "type": "Application Updated", "title": f"Application {app.application_ref} updated", "details": f"Status: {app.status}"})
        if app.signed_at:
            events.append({"when": app.signed_at, "type": "Signature", "title": f"Application {app.application_ref} signed", "details": app.signed_pdf_path or "Signed PDF generated"})
        if app.sign_token_created_at:
            events.append({"when": app.sign_token_created_at, "type": "Signing Link", "title": f"Signing link created for {app.application_ref}", "details": "Used" if app.sign_token_used_at else "Awaiting client signature"})
        if app.sign_token_used_at:
            events.append({"when": app.sign_token_used_at, "type": "Signing Link Used", "title": f"Signing link used for {app.application_ref}", "details": "Client opened/completed signing link"})

    for sig in signatures:
        events.append({"when": sig.signed_at, "type": "Client Signature", "title": sig.typed_name or "Client signed", "details": f"POPIA: {'Yes' if sig.consent_popia else 'No'}; Disclosure: {'Yes' if sig.consent_disclosure else 'No'}; OTP: {'Verified' if sig.otp_verified else 'Not verified'}"})

    for doc in fica_docs:
        events.append({"when": doc.uploaded_at, "type": "FICA Document", "title": doc.document_type.replace('_', ' ').title(), "details": f"{doc.status}: {doc.original_filename}"})

    events = sorted(events, key=lambda e: e["when"] or datetime.min, reverse=True)
    return render_template("recovery/timeline.html", p=p, calls=calls, applications=applications, sessions=sessions, signatures=signatures, fica_docs=fica_docs, events=events)



@recovery_bp.route("/suspense/rebuild", methods=["POST"])
@login_required
@permission_required("recovery.import")
def rebuild_suspense():
    base = scope_by_branch(LapsedPolicy.query, LapsedPolicy, agent_col=LapsedPolicy.assigned_agent_id)
    moved = _move_missing_policies_to_suspense(base)
    flash(f"Suspense check complete. {moved} policy/client record(s) moved to Suspense.", "success" if moved else "info")
    return redirect(url_for("recovery.suspense"))

@recovery_bp.route("/suspense")
@login_required
@permission_required("recovery.view")
def suspense():
    """Branch-specific suspense list for imported policies with missing contact details."""
    if not (can_view_all_branches() or is_branch_manager()):
        abort(403)
    selected_branch = (request.args.get("branch") or "").strip()
    query = scope_by_branch(
        LapsedPolicy.query.filter(LapsedPolicy.recovery_status == SUSPENSE_STATUS),
        LapsedPolicy,
        selected_branch=selected_branch,
        allow_agent_fallback=False,
    )
    policies = query.order_by(LapsedPolicy.branch.asc().nullslast(), LapsedPolicy.imported_at.desc()).all()
    for p in policies:
        if not getattr(p, "company_name", None):
            p.company_name = p.franchise or p.branch

    branch_groups = {}
    for p in policies:
        branch = (p.branch or p.franchise or "Unknown Branch").strip()
        branch_groups.setdefault(branch, []).append(p)

    branch_query = db.session.query(LapsedPolicy.branch).filter(
        LapsedPolicy.recovery_status == SUSPENSE_STATUS,
        LapsedPolicy.branch.isnot(None),
        LapsedPolicy.branch != ""
    )
    if not can_view_all_branches() and user_branch():
        branch_query = branch_query.filter(LapsedPolicy.branch == user_branch())
    branches = [r[0] for r in branch_query.distinct().order_by(LapsedPolicy.branch.asc()).all()]

    return render_template("recovery/suspense.html", branch_groups=branch_groups, branches=branches, selected_branch=selected_branch)


@recovery_bp.route("/import", methods=["POST"])
@login_required
@permission_required("recovery.import")
def import_lapsed():
    file = request.files.get("file")
    if not file:
        flash("Please choose an Excel file", "danger")
        return redirect(url_for("recovery.queue"))
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    count = 0
    suspense_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        policy_number = _row_value(data, ["Policy_Number", "Policy Number", "PolicyNumber"])
        if not policy_number:
            continue

        missing_fields, id_number, contact_number, email_address = _missing_contact_fields(data)
        branch = _row_value(data, ["CollectedatBranch", "Collected at Branch", "Branch", "Franchise"])
        company_name = _row_value(data, COMPANY_HEADERS) or branch or _row_value(data, ["Franchise"])
        comments = data.get("Comments")
        recovery_status = "Imported"
        next_action = date.today()
        assigned_agent = current_user.id

        if missing_fields:
            recovery_status = SUSPENSE_STATUS
            next_action = None
            assigned_agent = None
            suspense_count += 1
            comments = _append_comment(
                comments,
                "SUSPENSE: Missing " + ", ".join(missing_fields) + ". Client cannot be contacted until business client/branch fixes these policy details."
            )
        else:
            count += 1

        lp = LapsedPolicy(
            franchise=company_name, company_name=company_name, id_number=id_number, email_address=email_address, suspense_reason=", ".join(missing_fields), member_id=str(data.get("Member_ID") or ""), policy_number=str(policy_number),
            surname=data.get("Surname"), initials=data.get("Initials"), cell_number=contact_number or str(data.get("Cell_Number") or ""),
            home_tel=str(data.get("home_tel") or ""), address=data.get("Address"), premium_due=data.get("PremiumDue") or 0,
            total=data.get("Total") or 0, payment_method=data.get("PaymentMethod"), branch=branch,
            comments=comments, assigned_agent_id=assigned_agent, recovery_status=recovery_status, next_action_date=next_action
        )
        db.session.add(lp)
    db.session.commit()
    if suspense_count:
        flash(f"Import complete. Call queue: {count}. Suspense: {suspense_count} clients missing ID number, contact number or email address.", "warning")
        return redirect(url_for("recovery.suspense"))
    flash(f"Imported {count} lapsed policies", "success")
    return redirect(url_for("recovery.queue"))


@recovery_bp.route("/<int:policy_id>/call", methods=["GET", "POST"])
@login_required
@permission_required("recovery.call")
def log_call(policy_id):
    p = LapsedPolicy.query.get_or_404(policy_id)
    ensure_branch_access(p, agent_attr="assigned_agent_id")
    outcomes = CALL_OUTCOMES
    previous_calls = RecoveryCallLog.query.filter_by(lapsed_policy_id=p.id).order_by(RecoveryCallLog.created_at.desc()).limit(10).all()
    if request.method == "POST":
        outcome = request.form["outcome"]
        follow = request.form.get("follow_up_date") or None
        notes = (request.form.get("notes") or "").strip()

        if outcome in CALLBACK_OUTCOMES and not follow:
            follow = (date.today() + timedelta(days=1)).isoformat()

        next_action = date.fromisoformat(follow) if follow else None
        if outcome in {"Interested", "Application Started", "Signature Sent", "FICA Outstanding", "QA Review"} and not next_action:
            next_action = date.today()

        log = RecoveryCallLog(
            lapsed_policy_id=p.id,
            agent_id=current_user.id,
            outcome=outcome,
            notes=notes,
            follow_up_date=follow,
            next_action_date=next_action
        )

        p.recovery_status = OUTCOME_TO_STATUS.get(outcome, "Called")
        p.assigned_agent_id = current_user.id

        if outcome in ["Wants Reinstatement", "Wants New Policy"]:
            p.next_action_date = date.today()
            db.session.add(log)
            db.session.commit()
            app_type = "reinstatement" if outcome == "Wants Reinstatement" else "new"
            flash("Call logged. Start the full application process below. No joining fee will apply.", "success")
            return redirect(url_for("recovery.start_script", policy_id=p.id, app_type=app_type))

        p.next_action_date = next_action
        db.session.add(log)
        db.session.commit()
        flash("Call logged", "success")
        return redirect(url_for("recovery.queue"))

    return render_template("recovery/log_call.html", p=p, outcomes=outcomes, previous_calls=previous_calls)




SCRIPT_STEPS = [{'id': 1,
  'title': 'Confirm Client Identity',
  'qa': 'SECTION 1: INTRODUCTION & DISCLOSURE',
  'block_on_no': True,
  'script': 'Good day.\n\nAm I speaking with [Client Name]?',
  'question': 'Client confirmed you are speaking to the correct person?'},
 {'id': 2,
  'title': 'Agent Introduction',
  'qa': 'SECTION 1: INTRODUCTION & DISCLOSURE',
  'block_on_no': False,
  'script': "Thank you.\n\nMy name is [Agent Name].\n\nI am calling from Martin's Funerals.",
  'question': "Agent introduced themselves and Martin's Funerals?"},
 {'id': 3,
  'title': 'Call Recording Consent',
  'qa': 'SECTION 1: INTRODUCTION & DISCLOSURE',
  'block_on_no': True,
  'script': 'This call is recorded for quality, training and regulatory purposes.\n'
            '\n'
            'Do I have your consent to continue?',
  'question': 'Client gave consent for the recorded call to continue?'},
 {'id': 4,
  'title': 'Company Disclosure',
  'qa': 'SECTION 1: INTRODUCTION & DISCLOSURE',
  'block_on_no': False,
  'script': "Martin's Funerals offers funeral cover products designed to assist families with funeral expenses and "
            'related benefits.',
  'question': 'Company and product purpose disclosed?'},
 {'id': 5,
  'title': 'Age and Decision-Making Confirmation',
  'qa': 'SECTION 1: INTRODUCTION & DISCLOSURE',
  'block_on_no': True,
  'script': 'Before we continue, may I confirm that you are over the age of 18 and able to make financial decisions '
            'for yourself?',
  'question': 'Client confirmed they are over 18 and able to make financial decisions?'},
 {'id': 6,
  'title': 'Permission to Proceed',
  'qa': 'SECTION 1: INTRODUCTION & DISCLOSURE',
  'block_on_no': True,
  'script': 'Thank you.\n'
            '\n'
            'The purpose of this call is to explain our funeral cover options and determine whether they may be '
            'suitable for your needs.\n'
            '\n'
            'The call should take approximately 10 to 15 minutes.\n'
            '\n'
            'Do I have your permission to continue?',
  'question': 'Client gave permission to continue?'},
 {'id': 7,
  'title': 'Needs Analysis Introduction',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'To ensure I recommend the correct option, I would like to ask a few questions.',
  'question': 'Needs analysis started?'},
 {'id': 8,
  'title': 'Existing Funeral Cover',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'Do you currently have funeral cover?',
  'question': 'Client answered whether they currently have funeral cover?'},
 {'id': 9,
  'title': 'Who Needs Cover',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'Who would you like to cover?\n'
            '\n'
            'Please confirm if it is:\n'
            '\n'
            '- Yourself only\n'
            '- Yourself and spouse\n'
            '- Children\n'
            '- Extended family members',
  'question': 'Agent identified who the client wants to cover?'},
 {'id': 10,
  'title': 'Number of Lives',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'How many people would you like to include on the policy?',
  'question': 'Number of people to be covered was discussed?'},
 {'id': 11,
  'title': 'Cover Level',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'What level of funeral cover are you looking for?',
  'question': 'Cover level requirement was discussed?'},
 {'id': 12,
  'title': 'Affordability',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'Are affordability and monthly premium important considerations for you?',
  'question': 'Affordability was discussed?'},
 {'id': 13,
  'title': 'Recent Cover Elsewhere',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'Have any proposed insured persons recently taken out funeral cover elsewhere?',
  'question': 'Replacement/recent cover question was asked?'},
 {'id': 14,
  'title': 'Product Recommendation',
  'qa': 'SECTION 2: NEEDS ANALYSIS',
  'block_on_no': False,
  'script': 'Based on your responses, I believe the following option may be suitable.',
  'question': 'Suitable product recommendation was made?'},
 {'id': 15,
  'title': 'Product Explanation',
  'qa': 'SECTION 3: PRODUCT EXPLANATION',
  'block_on_no': True,
  'script': 'Our funeral cover provides financial assistance upon the death of an insured person.\n'
            '\n'
            'The policy offers funeral cover for eligible insured persons.\n'
            '\n'
            'A lump-sum benefit is paid to the nominated beneficiary.\n'
            '\n'
            'The cover depends on the selected plan.',
  'question': 'Core product explanation completed?'},
 {'id': 16,
  'title': 'Additional Benefits',
  'qa': 'SECTION 3: PRODUCT EXPLANATION',
  'block_on_no': False,
  'script': 'Additional benefits may apply where included in the specific product.\n'
            '\n'
            'These may include repatriation, grocery benefits, airtime benefits, or other value-added services.\n'
            '\n'
            'I will explain only the benefits included in the selected plan.',
  'question': 'Additional benefits were explained only where applicable?'},
 {'id': 17,
  'title': 'Selected Plan Details',
  'qa': 'SECTION 3: PRODUCT EXPLANATION',
  'block_on_no': True,
  'script': 'I will now explain the selected plan.\n'
            '\n'
            'Please confirm the plan name, main member cover, spouse cover, child cover, extended family cover, and '
            'any other selected benefits.',
  'question': 'Selected plan, cover amounts and benefits were explained?'},
 {'id': 18,
  'title': 'Monthly Premium',
  'qa': 'SECTION 4: PREMIUM DISCLOSURE',
  'block_on_no': True,
  'script': 'The monthly premium for this policy will be explained to you.\n'
            '\n'
            'For this premium, you will receive the benefits I have just explained.',
  'question': 'Monthly premium was disclosed?'},
 {'id': 19,
  'title': 'Payment Frequency and Non-Payment',
  'qa': 'SECTION 4: PREMIUM DISCLOSURE',
  'block_on_no': True,
  'script': 'Premiums are payable monthly according to the selected payment method.\n'
            '\n'
            'Please note that if premiums are not paid, benefits may be reduced, suspended, or the policy may lapse '
            'according to the policy terms and conditions.',
  'question': 'Payment frequency and non-payment consequences were explained?'},
 {'id': 20,
  'title': 'Waiting Periods',
  'qa': 'SECTION 5: WAITING PERIODS & EXCLUSIONS',
  'block_on_no': True,
  'script': 'It is important that I explain the waiting periods.\n'
            '\n'
            'A waiting period of up to 6 months may apply for death due to natural causes.\n'
            '\n'
            'Cover for accidental death may commence immediately, subject to policy terms.',
  'question': 'Waiting periods and accidental death rules were explained?'},
 {'id': 21,
  'title': 'Exclusions',
  'qa': 'SECTION 5: WAITING PERIODS & EXCLUSIONS',
  'block_on_no': True,
  'script': 'Certain limitations or waiting periods may apply where permitted by law and policy rules.\n'
            '\n'
            'A suicide exclusion period may apply as specified in the policy wording.\n'
            '\n'
            'Claims may be declined if false, incomplete, or misleading information is provided during the application '
            'process.',
  'question': 'Exclusions, suicide period and misrepresentation risk were explained?'},
 {'id': 22,
  'title': 'Waiting Period Understanding',
  'qa': 'SECTION 5: WAITING PERIODS & EXCLUSIONS',
  'block_on_no': True,
  'script': 'Do you understand these waiting periods and exclusions?',
  'question': 'Client confirmed understanding of waiting periods and exclusions?'},
 {'id': 23,
  'title': 'Terms and Conditions',
  'qa': 'SECTION 6: TERMS & CONDITIONS',
  'block_on_no': True,
  'script': 'I would also like to disclose the following important terms and conditions.\n'
            '\n'
            "The policy is subject to the insurer's policy wording.\n"
            '\n'
            'Premiums may change according to policy provisions.\n'
            '\n'
            'You have the right to receive all policy documentation.\n'
            '\n'
            'You may cancel the policy according to applicable regulations and policy terms.\n'
            '\n'
            'Claims must be supported by the required documentation.\n'
            '\n'
            'Benefits will only be paid if all policy requirements are met.',
  'question': 'Terms and conditions were disclosed?'},
 {'id': 24,
  'title': 'Terms Acceptance',
  'qa': 'SECTION 6: TERMS & CONDITIONS',
  'block_on_no': True,
  'script': 'Do you understand and accept these terms and conditions?',
  'question': 'Client accepted the terms and conditions?'},
 {'id': 25,
  'title': 'Confirm Client Details',
  'qa': 'SECTION 7: CLIENT DETAILS VERIFICATION',
  'block_on_no': True,
  'script': 'I will now confirm your details.\n'
            '\n'
            'Please confirm your full names, surname, South African ID number, date of birth, physical address, postal '
            'address if applicable, contact number, and email address.',
  'question': 'Client personal and contact details were confirmed?'},
 {'id': 26,
  'title': 'Confirm Beneficiary Details',
  'qa': 'SECTION 7: CLIENT DETAILS VERIFICATION',
  'block_on_no': True,
  'script': 'Please confirm the beneficiary name, beneficiary contact number, and relationship to the beneficiary.\n'
            '\n'
            'Are all these details correct and complete?',
  'question': 'Beneficiary details were confirmed as correct?'},
 {'id': 27,
  'title': 'Debit Order Details',
  'qa': 'SECTION 8: DEBIT ORDER AUTHORISATION',
  'block_on_no': False,
  'script': 'If premium collection is done via debit order, please confirm the account holder name, bank name, account '
            'number, branch code, account type and preferred debit date.',
  'question': 'Banking details were verified where Debit Order applies?'},
 {'id': 28,
  'title': 'Debit Order Authorisation',
  'qa': 'SECTION 8: DEBIT ORDER AUTHORISATION',
  'block_on_no': False,
  'script': "Do you authorise Martin's Funerals and/or its authorised collection partner to deduct the monthly premium "
            'from your bank account on the agreed debit date?\n'
            '\n'
            'Please answer yes or no.',
  'question': 'Debit order authorisation obtained where applicable?'},
 {'id': 29,
  'title': 'POPIA Consent',
  'qa': 'SECTION 9: POPIA CONSENT',
  'block_on_no': True,
  'script': "In accordance with POPIA, Martin's Funerals requires your consent to collect, process, store, and use "
            'your personal information for:\n'
            '\n'
            '- Policy administration\n'
            '- Underwriting\n'
            '- Claims processing\n'
            '- Regulatory compliance\n'
            '- Customer service and communication\n'
            '\n'
            'Your information will be handled confidentially and in accordance with applicable legislation.\n'
            '\n'
            'Do you consent to the processing of your personal information for these purposes?',
  'question': 'POPIA consent received?'},
 {'id': 30,
  'title': 'Confirmation of Understanding',
  'qa': 'SECTION 10: CLOSING',
  'block_on_no': True,
  'script': 'Before we proceed, I would like to confirm that I have explained:\n'
            '\n'
            '- Policy benefits\n'
            '- Premiums\n'
            '- Waiting periods and exclusions\n'
            '- Terms and conditions\n'
            '- Your personal information\n'
            '- Debit order authorisation, where applicable\n'
            '- Your POPIA rights\n'
            '\n'
            'Do you confirm that you understand the information provided and wish to proceed with the policy '
            'application?',
  'question': 'Client understands and wants to proceed with the application?'},
 {'id': 31,
  'title': 'Closing and Next Steps',
  'qa': 'SECTION 10: CLOSING',
  'block_on_no': False,
  'script': "Thank you for choosing Martin's Funerals.\n"
            '\n'
            'Your application will now be submitted for processing.\n'
            '\n'
            'You will receive your policy documents and confirmation via SMS, email, WhatsApp, or your preferred '
            'communication method.\n'
            '\n'
            'Please review your documents carefully and contact us immediately if any information is incorrect.\n'
            '\n'
            "Thank you for your time today and welcome to Martin's Funerals.\n"
            '\n'
            'Have a wonderful day.',
  'question': 'Closing completed professionally?'}]

QA_SECTIONS = [
    ("SECTION 1: INTRODUCTION & DISCLOSURE", 10),
    ("SECTION 2: NEEDS ANALYSIS", 10),
    ("SECTION 3: PRODUCT EXPLANATION", 15),
    ("SECTION 4: PREMIUM DISCLOSURE", 10),
    ("SECTION 5: WAITING PERIODS & EXCLUSIONS", 15),
    ("SECTION 6: TERMS & CONDITIONS", 10),
    ("SECTION 7: CLIENT DETAILS VERIFICATION", 10),
    ("SECTION 8: DEBIT ORDER AUTHORISATION", 10),
    ("SECTION 9: POPIA CONSENT", 5),
    ("SECTION 10: CLOSING", 5),
]


def _answer_value(session, key, default=None):
    answers = _script_answers(session)
    for item in answers.values():
        if isinstance(item, dict) and item.get(key) not in (None, ""):
            return item.get(key)
    return default


def _client_surname(session):
    if session and session.lapsed_policy and session.lapsed_policy.surname:
        return session.lapsed_policy.surname
    parts = (session.client_name or "").split()
    return parts[-1] if parts else "Client"


def _client_title(session):
    # Use captured title first. If no title exists, infer a respectful salutation from the SA ID gender digits where possible.
    title = _answer_value(session, "client_title")
    if title:
        return str(title).strip()
    id_number = _script_client_id_number(session)
    try:
        if id_number and len(id_number) >= 10 and id_number[:10].isdigit():
            return "Mrs" if int(id_number[6:10]) < 5000 else "Mr"
    except Exception:
        pass
    return "Mr/Mrs"

def _client_display(session):
    surname = _client_surname(session)
    initials = ""
    if session and session.lapsed_policy:
        initials = session.lapsed_policy.initials or ""
    return f"{_client_title(session)} {surname}".strip() or surname

def _agent_display_name():
    return (getattr(current_user, "agent_name", None) or getattr(current_user, "name", None) or getattr(current_user, "email", None) or "Agent").strip()

def _selected_additional_benefits(session):
    return _answer_value(session, "additional_benefits", "") or ""


def _script_client_id_number(session):
    """Return the best available client ID number for script/product filtering."""
    saved = _answer_value(session, "client_id_number")
    if saved:
        return str(saved).strip()
    if session and session.lapsed_policy:
        candidates = [
            session.lapsed_policy.member_id,
            session.lapsed_policy.policy_number,
            session.lapsed_policy.comments,
        ]
        for value in candidates:
            match = re.search(r"\b\d{13}\b", str(value or ""))
            if match:
                return match.group(0)
    return ""


def _script_client_age(session):
    age = _answer_value(session, "client_age")
    try:
        if str(age or "").strip():
            return int(age)
    except Exception:
        pass
    id_number = _script_client_id_number(session)
    try:
        dob = dob_from_sa_id(id_number) if id_number else None
        return age_from_dob(dob) if dob else None
    except Exception:
        return None


def _product_label(product):
    if not product:
        return ""
    return f"{product.product_name or ''} / {product.plan_name or ''}".strip(" / ")


def _selected_script_product(session):
    prod_id = _answer_value(session, "product_id")
    try:
        return PolicyProduct.query.get(int(prod_id)) if prod_id else None
    except Exception:
        return None


def _coverage_filter(products, coverage_choice):
    choice = (coverage_choice or "").lower()
    if not choice:
        return products
    filtered = []
    for prod in products:
        text = f"{prod.product_name or ''} {prod.plan_name or ''}".lower()
        if choice == "myself_only":
            if any(k in text for k in ["single", "individual", "member", "product"]):
                filtered.append(prod)
        elif choice in {"myself_spouse", "myself_spouse_children", "family"}:
            if any(k in text for k in ["family", "spouse", "child", "children"]):
                filtered.append(prod)
        elif choice == "extended_family":
            if any(k in text for k in ["extended", "family"]):
                filtered.append(prod)
    return filtered or products


def _eligible_products_for_script(session):
    age = _script_client_age(session)
    coverage = _answer_value(session, "coverage_choice")
    products = PolicyProduct.query.filter_by(active=True).order_by(PolicyProduct.product_name, PolicyProduct.plan_name).all()
    eligible = []
    for prod in products:
        if age is not None:
            if prod.min_age is not None and age < int(prod.min_age):
                continue
            if prod.max_age is not None and age > int(prod.max_age):
                continue
        eligible.append(prod)
    return _coverage_filter(eligible, coverage)

def _selected_script_beneficiary(session):
    return {
        "name": _answer_value(session, "beneficiary_name", "") or "",
        "contact": _answer_value(session, "beneficiary_contact", "") or "",
        "relationship": _answer_value(session, "beneficiary_relationship", "") or "",
    }


def _selected_script_bank_details(session):
    return {
        "account_holder": _answer_value(session, "account_holder", "") or "",
        "bank_name": _answer_value(session, "bank_name", "") or "",
        "account_number": _answer_value(session, "account_number", "") or "",
        "branch_code": _answer_value(session, "branch_code", "") or "",
        "account_type": _answer_value(session, "account_type", "") or "",
        "debit_day": _answer_value(session, "debit_day", "") or "",
    }


def _existing_member_summary(session):
    p = session.lapsed_policy if session else None
    if not p:
        return []
    rows = []
    if p.surname or p.initials:
        rows.append({"role": "Principal Member", "name": f"{p.initials or ''} {p.surname or ''}".strip(), "id": _script_client_id_number(session), "contact": p.cell_number or p.home_tel or ""})
    # Current import file does not always include spouse/children/extended rows. Keep the structure so it can display when those fields are later imported.
    return rows


def _script_answers_for_application(session):
    return {
        "client_id_number": _script_client_id_number(session),
        "client_age": _script_client_age(session),
        "coverage_choice": _answer_value(session, "coverage_choice", "") or "",
        "additional_benefits": _selected_additional_benefits(session),
        "beneficiary": _selected_script_beneficiary(session),
        "bank": _selected_script_bank_details(session),
    }


def _format_money(value):
    try:
        return f"R {float(value or 0):,.2f}"
    except Exception:
        return "R 0.00"


def _script_text_for_display(session, step):
    text = step.get("script", "")
    product = _selected_script_product(session)
    payment = _answer_value(session, "payment_method", "the selected payment method")
    benefits = _selected_additional_benefits(session)
    display = _client_display(session)

    text = text.replace("[Client Name]", display or "Client")
    text = text.replace("[Client Surname]", _client_surname(session) or "Client")
    text = text.replace("[Agent Name]", _agent_display_name())
    text = text.replace("the selected payment method", payment)

    if product and step.get("id") == 18:
        text = text.replace("The monthly premium for this policy will be explained to you.", f"The monthly premium for this policy will be {_format_money(product.monthly_premium)}.")

    if product and step.get("id") in {17, 18, 24, 25}:
        details = [
            "Selected policy details:",
            f"Plan: {_product_label(product)}",
            f"Cover Amount: {_format_money(product.cover_amount)}",
            f"Monthly Premium: {_format_money(product.monthly_premium)}",
            f"Waiting Period: {product.waiting_period_months or 0} months",
        ]
        if benefits:
            details.append(f"Additional Benefits: {benefits}")
        text += "\n\n" + "\n".join(details)

    return text

def _application_salutation(app_obj):
    title = (getattr(app_obj, "title", "") or "").strip()
    surname = (getattr(app_obj, "surname", "") or getattr(app_obj, "first_names", "") or "Client").strip()
    if title:
        return f"{title} {surname}"
    return surname

def _send_script_selected_signing_link(app_obj, delivery_method):
    ok, errors = assert_application_rules(app_obj)
    if not ok:
        return None, False, errors
    token = secrets.token_urlsafe(32)
    app_obj.sign_token = token
    app_obj.sign_token_created_at = datetime.utcnow()
    app_obj.sign_token_used_at = None
    app_obj.sign_token_revoked = False
    folder = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(folder, exist_ok=True)
    preview_pdf = os.path.join(folder, f"review_application_{app_obj.id}.pdf")
    popia_pdf = os.path.join(folder, f"popia_consent_{app_obj.id}.pdf")
    disclosure_pdf = os.path.join(folder, f"policy_disclosure_{app_obj.id}.pdf")
    fica_pdf = os.path.join(folder, f"fica_verification_{app_obj.id}.pdf")
    generate_application_pdf(app_obj, preview_pdf)
    generate_popia_pdf(app_obj, popia_pdf)
    generate_disclosure_pdf(app_obj, disclosure_pdf)
    generate_fica_pdf(app_obj, fica_pdf)
    app_obj.popia_pdf_path = popia_pdf
    app_obj.disclosure_pdf_path = disclosure_pdf
    base_url = current_app.config.get("BASE_URL") or request.url_root.rstrip("/")
    link = f"{base_url}{url_for('signing.sign_application', token=token)}"
    body = (
        f"Dear {_application_salutation(app_obj)},\n\n"
        "Please open this secure Martin's Funerals link to upload your FICA documents and sign your application documents:\n\n"
        f"{link}\n\n"
        "You will need your ID number to unlock the page.\n\n"
        "No documents are attached. Your documents are available only inside the secure signing link."
    )
    method = (delivery_method or "email").lower()
    sent = False
    if method in {"email", "sms_email", "whatsapp_email"} and app_obj.email:
        sent = send_email(app_obj.email, "Your Martin's Funerals secure signing link", body, []) or sent
    if method in {"whatsapp", "whatsapp_email"} and app_obj.cell_number:
        sent = send_whatsapp_message(app_obj.cell_number, body) or sent
    if method == "sms":
        current_app.logger.info("SMS selected for signing link, but no SMS provider is configured. Link: %s", link)
    app_obj.status = "Signing Link Sent" if sent else "Signing Link Prepared"
    db.session.commit()
    return link, sent, []


SCRIPT_CONFIG_FILE = "telesales_script_steps.json"


def _script_config_path():
    instance_path = getattr(current_app, "instance_path", None) or current_app.root_path
    os.makedirs(instance_path, exist_ok=True)
    return os.path.join(instance_path, SCRIPT_CONFIG_FILE)


def _default_script_steps():
    # Make a safe copy so admin edits never mutate the in-code defaults.
    return json.loads(json.dumps(SCRIPT_STEPS))


def _load_script_steps():
    """Load admin-edited wording while keeping the original flow/order/QA mapping."""
    steps = _default_script_steps()
    path = _script_config_path()
    if not os.path.exists(path):
        return steps
    try:
        with open(path, "r", encoding="utf-8") as f:
            edited = json.load(f)
        edited_by_id = {int(item.get("id")): item for item in edited if item.get("id") is not None}
        for step in steps:
            edit = edited_by_id.get(int(step["id"]))
            if not edit:
                continue
            # Admin may change wording/questions, but not the compliance flow id/order/QA section.
            for field in ("title", "script", "question", "block_on_no"):
                if field in edit:
                    step[field] = edit[field]
        return steps
    except Exception as exc:
        current_app.logger.exception("Could not load telesales script config: %s", exc)
        return steps


def _save_script_steps_from_form():
    steps = _load_script_steps()
    edited = []
    for step in steps:
        sid = str(step["id"])
        edited.append({
            "id": step["id"],
            "title": request.form.get(f"title_{sid}", step["title"]).strip(),
            "script": request.form.get(f"script_{sid}", step["script"]).strip(),
            "question": request.form.get(f"question_{sid}", step["question"]).strip(),
            "block_on_no": request.form.get(f"block_on_no_{sid}") == "on",
        })
    with open(_script_config_path(), "w", encoding="utf-8") as f:
        json.dump(edited, f, ensure_ascii=False, indent=2)


def _current_script_steps():
    return _load_script_steps()


def _role_name():
    return str(getattr(getattr(current_user, "role", None), "name", "") or "").lower()


def _can_manage_scripts():
    role = _role_name()
    return role in {"admin", "branch manager", "branch_manager", "manager"}


def _script_step(step_id):
    for step in _current_script_steps():
        if step["id"] == step_id:
            return step
    return None


def _script_answers(session):
    try:
        return json.loads(session.answers_json or "{}")
    except Exception:
        return {}


def _script_score(answers):
    # QA checklist cross-reference: each section passes if all linked required script answers are Yes/NA.
    section_ok = {name: True for name, _ in QA_SECTIONS}
    for step in _current_script_steps():
        answer = (answers.get(str(step["id"]), {}) or {}).get("answer")
        if answer == "no" and step.get("block_on_no", False):
            section_ok[step["qa"]] = False
    total = sum(points for name, points in QA_SECTIONS if section_ok.get(name, True))
    return total, "PASS" if total >= 90 else "FAIL"


@recovery_bp.route("/<int:policy_id>/script/start")
@login_required
@permission_required("recovery.call")
def start_script(policy_id):
    p = LapsedPolicy.query.get_or_404(policy_id)
    ensure_branch_access(p, agent_attr="assigned_agent_id")
    app_type = request.args.get("app_type", "new")
    client_name = f"{p.initials or ''} {p.surname or ''}".strip()
    session = TelesalesScriptSession(
        lapsed_policy_id=p.id,
        agent_id=current_user.id,
        branch=p.branch or current_user.branch,
        client_name=client_name,
        client_cell=p.cell_number or p.home_tel,
        policy_number=p.policy_number,
        script_type=app_type,
        status="In Progress",
        current_step=1,
        answers_json="{}",
    )
    p.recovery_status = "Script In Progress"
    db.session.add(session)
    db.session.commit()
    return redirect(url_for("recovery.script_step", session_id=session.id))


@recovery_bp.route("/script/<int:session_id>", methods=["GET", "POST"])
@login_required
@permission_required("recovery.call")
def script_step(session_id):
    session = TelesalesScriptSession.query.get_or_404(session_id)
    ensure_branch_access(session, agent_attr="agent_id")
    if session.agent_id != current_user.id and not _can_manage_scripts():
        abort(403)
    step = _script_step(session.current_step)
    if not step:
        return redirect(url_for("recovery.script_complete", session_id=session.id))
    # Only ask number of lives when extended cover was selected. Otherwise skip this step automatically.
    if request.method == "GET" and step.get("id") == 10 and (_answer_value(session, "coverage_choice") != "extended_family"):
        session.current_step = 11
        db.session.commit()
        return redirect(url_for("recovery.script_step", session_id=session.id))
    if request.method == "POST":
        if request.form.get("go_to_step"):
            try:
                session.current_step = int(request.form.get("go_to_step"))
                db.session.commit()
                return redirect(url_for("recovery.script_step", session_id=session.id))
            except Exception:
                pass
        answer = request.form.get("answer")
        note = request.form.get("note", "")
        extra = {}
        if step["id"] == 9:
            extra["coverage_choice"] = request.form.get("coverage_choice") or answer
            extra["client_id_number"] = request.form.get("client_id_number") or _script_client_id_number(session)
            dob = dob_from_sa_id(extra["client_id_number"]) if extra.get("client_id_number") else None
            calculated_age = age_from_dob(dob) if dob else None
            extra["client_age"] = calculated_age or request.form.get("client_age") or _script_client_age(session)
            answer = extra["coverage_choice"] or answer
        if step["id"] == 11:
            if request.form.get("product_id"):
                extra["product_id"] = request.form.get("product_id")
                answer = "yes"
        if step["id"] == 16:
            extra["additional_benefits"] = request.form.get("additional_benefits") or ""
            answer = "yes"
        if step["id"] == 12 and answer == "no":
            # Premium too high: return agent to policy selection instead of continuing.
            answers = _script_answers(session)
            answers[str(step["id"])] = {"answer": answer, "note": note, "title": step["title"], "qa": step["qa"], "question": step["question"], "recorded_at": datetime.utcnow().isoformat(), **extra}
            session.answers_json = json.dumps(answers)
            session.current_step = 11
            db.session.commit()
            flash("Premium declined. Please select a different product in the client's age range.", "warning")
            return redirect(url_for("recovery.script_step", session_id=session.id))
        if step["id"] == 19:
            extra["payment_method"] = request.form.get("payment_method") or answer
            answer = extra["payment_method"] or answer
        if step["id"] == 26:
            extra["beneficiary_name"] = request.form.get("beneficiary_name") or ""
            extra["beneficiary_contact"] = request.form.get("beneficiary_contact") or ""
            extra["beneficiary_relationship"] = request.form.get("beneficiary_relationship") or ""
            answer = "yes"
        if step["id"] == 27:
            extra["account_holder"] = request.form.get("account_holder") or ""
            extra["bank_name"] = request.form.get("bank_name") or ""
            extra["account_number"] = request.form.get("account_number") or ""
            extra["branch_code"] = request.form.get("branch_code") or ""
            extra["account_type"] = request.form.get("account_type") or ""
            extra["debit_day"] = request.form.get("debit_day") or ""
            answer = "yes"
        if step["id"] == 31:
            extra["delivery_method"] = request.form.get("delivery_method") or answer
            answer = extra["delivery_method"] or answer
        answers = _script_answers(session)
        answers[str(step["id"])] = {"answer": answer, "note": note, "title": step["title"], "qa": step["qa"], "question": step["question"], "recorded_at": datetime.utcnow().isoformat(), **extra}
        session.answers_json = json.dumps(answers)
        if answer == "no" and step.get("block_on_no"):
            session.status = "Blocked"
            session.blocked_reason = f"Client answered No at step {step['id']}: {step['title']}"
            session.completed_at = datetime.utcnow()
            session.qa_score, session.qa_result = _script_score(answers)
            db.session.commit()
            _save_script_pdf(session)
            flash(session.blocked_reason, "danger")
            return redirect(url_for("recovery.script_complete", session_id=session.id))
        session.current_step += 1
        if session.current_step > len(_current_script_steps()):
            session.status = "Completed"
            session.completed_at = datetime.utcnow()
            session.qa_score, session.qa_result = _script_score(answers)
            db.session.commit()
            _save_script_pdf(session)
            return redirect(url_for("recovery.script_complete", session_id=session.id))
        db.session.commit()
        return redirect(url_for("recovery.script_step", session_id=session.id))
    total_steps = len(_current_script_steps())
    progress = int(((session.current_step - 1) / total_steps) * 100)
    products = _eligible_products_for_script(session) if step["id"] == 11 else []
    selected_product = _selected_script_product(session)
    spoken_text = _script_text_for_display(session, step)
    return render_template("recovery/script_step.html", session=session, step=step, total_steps=total_steps, progress=progress, products=products, selected_product=selected_product, spoken_text=spoken_text, client_age=_script_client_age(session), client_id_number=_script_client_id_number(session), selected_payment=_answer_value(session, "payment_method"), selected_delivery=_answer_value(session, "delivery_method"), selected_beneficiary=_selected_script_beneficiary(session), selected_bank=_selected_script_bank_details(session), existing_members=_existing_member_summary(session), script_payload=_script_answers_for_application(session))


def _save_script_pdf(session):
    folder = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"telesales_script_qa_{session.id}.pdf")
    generate_telesales_script_pdf(session, _current_script_steps(), QA_SECTIONS, path)
    session.pdf_path = path
    db.session.commit()


@recovery_bp.route("/script/<int:session_id>/complete")
@login_required
@permission_required("recovery.call")
def script_complete(session_id):
    session = TelesalesScriptSession.query.get_or_404(session_id)
    ensure_branch_access(session, agent_attr="agent_id")
    if session.agent_id != current_user.id and not _can_manage_scripts():
        abort(403)
    answers = _script_answers(session)
    if not session.pdf_path and session.status in {"Completed", "Blocked"}:
        _save_script_pdf(session)
    return render_template("recovery/script_complete.html", session=session, answers=answers)


@recovery_bp.route("/script/<int:session_id>/download")
@login_required
@permission_required("applications.view")
def download_script_pdf(session_id):
    if not _can_manage_scripts():
        abort(403)
    session = TelesalesScriptSession.query.get_or_404(session_id)
    ensure_branch_access(session, agent_attr="agent_id")
    if not session.pdf_path or not os.path.exists(session.pdf_path):
        _save_script_pdf(session)
    return send_file(session.pdf_path, as_attachment=False)


@recovery_bp.route("/scripts")
@login_required
@permission_required("applications.view")
def script_records():
    if not _can_manage_scripts():
        abort(403)
    sessions = scope_by_branch(TelesalesScriptSession.query, TelesalesScriptSession, agent_col=TelesalesScriptSession.agent_id).order_by(TelesalesScriptSession.created_at.desc()).limit(300).all()
    return render_template("recovery/script_records.html", sessions=sessions)




@recovery_bp.route("/scripts/admin/questions", methods=["GET", "POST"])
@login_required
@permission_required("applications.view")
def admin_script_questions():
    if _role_name() != "admin":
        abort(403)
    if request.method == "POST":
        _save_script_steps_from_form()
        flash("Telesales script wording updated. The compliance flow and QA sections were kept unchanged.", "success")
        return redirect(url_for("recovery.admin_script_questions"))
    return render_template("recovery/admin_script_questions.html", steps=_current_script_steps(), qa_sections=QA_SECTIONS)


@recovery_bp.route("/scripts/admin/questions/reset", methods=["POST"])
@login_required
@permission_required("applications.view")
def reset_script_questions():
    if _role_name() != "admin":
        abort(403)
    path = _script_config_path()
    if os.path.exists(path):
        os.remove(path)
    flash("Telesales script wording reset to the default version.", "success")
    return redirect(url_for("recovery.admin_script_questions"))

@recovery_bp.route("/<int:policy_id>/start-application", methods=["GET", "POST"])
@login_required
@permission_required("applications.create")
def start_application(policy_id):
    p = LapsedPolicy.query.get_or_404(policy_id)
    ensure_branch_access(p, agent_attr="assigned_agent_id")
    app_type = request.args.get("app_type", "reinstatement")
    script_id = request.args.get("script_id") or request.form.get("script_id")
    script_session = TelesalesScriptSession.query.get(script_id) if script_id else None
    if script_session:
        ensure_branch_access(script_session, agent_attr="agent_id")
    products = _eligible_products_for_script(script_session) if script_session else PolicyProduct.query.filter_by(active=True).order_by(PolicyProduct.product_name, PolicyProduct.plan_name).all()
    selected_product = _selected_script_product(script_session) if script_session else None
    selected_payment = _answer_value(script_session, "payment_method") if script_session else ""
    selected_delivery = _answer_value(script_session, "delivery_method") if script_session else ""
    script_payload = _script_answers_for_application(script_session) if script_session else {}

    if request.method == "POST":
        product_id = request.form.get("product_id") or (selected_product.id if selected_product else None)
        prod = PolicyProduct.query.get(product_id) if product_id else None
        if not prod:
            flash("Please select a policy product.", "danger")
            return redirect(url_for("recovery.start_script", policy_id=p.id, app_type=app_type))

        application_ref = "APP-" + datetime.now().strftime("%Y%m%d") + "-" + secrets.token_hex(3).upper()
        first_names = request.form.get("first_names") or p.initials or ""
        surname = request.form.get("surname") or p.surname or ""
        email = request.form.get("email") or getattr(script_session, "client_email", None) or ""
        cell = request.form.get("cell_number") or p.cell_number or p.home_tel or ""
        beneficiary = script_payload.get("beneficiary", {}) if script_payload else {}
        bank = script_payload.get("bank", {}) if script_payload else {}
        id_number = request.form.get("id_number") or script_payload.get("client_id_number") or ""

        label = "Reinstatement" if app_type == "reinstatement" else "Lapsed New Policy"
        product_text = ((prod.product_name or "") + " " + (prod.plan_name or "")).lower()
        form_template = "member_product" if ("member +" in product_text or ("product" in product_text and ("+" in product_text or "member" in product_text))) else "single_family"

        a = ClientApplication(
            application_ref=application_ref,
            product_id=prod.id,
            branch=p.branch or current_user.branch,
            agent_id=current_user.id,
            application_type=label,
            lapsed_policy_id=p.id,
            original_policy_number=p.policy_number,
            first_names=first_names,
            agent_name=current_user.name,
            agent_code="",
            surname=surname,
            id_number=id_number,
            cell_number=cell,
            email=email,
            address=request.form.get("address") or p.address,
            residential_address=request.form.get("address") or p.address,
            cover_amount=prod.cover_amount,
            monthly_premium=prod.monthly_premium,
            waiting_period=f"{prod.waiting_period_months} months",
            joining_fee=0,
            joining_fee_waived=True,
            joining_fee_waiver_reason="Lapsed policy recovery - no joining fee applies",
            status="Draft - Lapsed Recovery",
            form_template=form_template,
            payment_method=request.form.get("payment_method") or selected_payment,
            date_of_birth=dob_from_sa_id(id_number or ""),
            plan_choice="Member + Product" if form_template == "member_product" else "",
            total_payment=prod.monthly_premium,
            beneficiary_full_names=beneficiary.get("name") or request.form.get("beneficiary_full_names") or "",
            beneficiary_relationship=beneficiary.get("relationship") or request.form.get("beneficiary_relationship") or "",
            account_holder=bank.get("account_holder") or request.form.get("account_holder") or "",
            bank_name=bank.get("bank_name") or request.form.get("bank_name") or "",
            account_number=bank.get("account_number") or request.form.get("account_number") or "",
            branch_code=bank.get("branch_code") or request.form.get("branch_code") or "",
            account_type=bank.get("account_type") or request.form.get("account_type") or "",
            debit_day=bank.get("debit_day") or request.form.get("debit_day") or "",
            product_dependents_json=json.dumps(_existing_member_summary(script_session)) if script_session else "[]"
        )
        if script_id:
            script_session = TelesalesScriptSession.query.get(script_id)
            if script_session:
                ensure_branch_access(script_session, agent_attr="agent_id")
            if script_session:
                script_session.application = a
                script_session.client_email = email

        p.recovery_status = "Application Started"
        p.next_action_date = None

        db.session.add(a)
        db.session.commit()

        delivery_method = request.form.get("delivery_method") or selected_delivery
        if delivery_method:
            link, sent, send_errors = _send_script_selected_signing_link(a, delivery_method)
            if send_errors:
                for error in send_errors:
                    flash(error, "danger")
                flash("Application created, but policy/FICA validation blocked delivery. No email, SMS or WhatsApp link was sent. Fix the product or required FICA details before sending the signing link.", "danger")
            elif sent:
                flash("Application created and signing link sent using the selected delivery method.", "success")
            else:
                flash(f"Application created. Signing link prepared but not sent automatically. Link: {link}", "warning")
        else:
            flash("Application created from lapsed policy. Joining fee waived. Continue with signing process and welcome pack.", "success")
        if delivery_method:
            flash("Returning to the recovery queue so the next client can be contacted.", "info")
            return redirect(url_for("recovery.queue"))
        return redirect(url_for("applications.view_application", app_id=a.id))

    return render_template("recovery/start_application.html", p=p, products=products, app_type=app_type, script_id=script_id, selected_product=selected_product, selected_payment=selected_payment, selected_delivery=selected_delivery, script_payload=script_payload, existing_members=_existing_member_summary(script_session) if script_session else [])
